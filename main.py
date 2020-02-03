from PyQt5 import uic
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys
import time
from datetime import datetime
from openpyxl import Workbook
import os
import requests

class AnkiGUI(QMainWindow):

	def __init__(self):
		super(AnkiGUI, self).__init__()
		uic.loadUi('view.ui', self)
		
		self.startBtn = self.findChild(QPushButton, 'startBtn') # Find the button
		self.startBtn.clicked.connect(self.startScrape)

		self.resumeBtn = self.findChild(QPushButton, 'resumeBtn') 
		self.resumeBtn.clicked.connect(self.resumeScrape)

		self.exportBtn = self.findChild(QPushButton, 'exportBtn') # Find the button
		self.exportBtn.clicked.connect(self.exportExcel)
		
		self.idsEdit = self.findChild(QTextEdit, 'idsEdit')
		self.idsEdit.setPlaceholderText("Please copy ID list and paste here...")

		self.idListView = self.findChild(QTableView, "idListView")
		self.resultModel = QStandardItemModel()
		self.resultModel.setHorizontalHeaderLabels(['ID', 'Status', 'Name'])
		self.idListView.setModel(self.resultModel)
		id_list_view_width = self.idListView.width()
		self.idListView.setColumnWidth(0, id_list_view_width * 1/ 12)
		self.idListView.setColumnWidth(1, id_list_view_width * 1/ 12)
		self.idListView.setColumnWidth(2, id_list_view_width * 9/ 12)
		self.idListView.showGrid()

		self.failedListView = self.findChild(QTableView, "failedListView")
		self.failedResultModel = QStandardItemModel()
		self.failedResultModel.setHorizontalHeaderLabels(['ID', 'Status', 'Message'])
		self.failedListView.setModel(self.failedResultModel)
		failedListView_view_width = self.failedListView.width()
		self.failedListView.setColumnWidth(0, failedListView_view_width * 1/ 12)
		self.failedListView.setColumnWidth(1, failedListView_view_width * 1/ 12)
		self.failedListView.setColumnWidth(2, failedListView_view_width * 9/ 12)
		self.failedListView.showGrid()
		
		self.progressBar = self.findChild(QProgressBar, 'progressBar')
		self.progressBar.setValue(0)

		self.idList = []
		self.fail_id_list = []
		self.total_id_count = 0
		self.success_id_count = 0
		self.passed_id_count = 0
		self.result_list = []
		self.show()

	def startScrape(self):
		self.get_ids_from_edit()
		if len(self.idList):

			self.startBtn.setEnabled(False)
			self.resumeBtn.setEnabled(False)
			self.exportBtn.setEnabled(False)
			self.progressBar.setValue(0)

			self.result_list.clear()
			c = 0

			self.passed_id_count = 0
			self.success_id_count = 0
			self.fail_id_count = 0

			self.resultModel.clear()
			self.failedResultModel.clear()
			self.fail_id_list.clear()

			self.resultModel.setHorizontalHeaderLabels(['ID', 'Status', 'Name'])
			id_list_view_width = self.idListView.width()
			self.idListView.setColumnWidth(0, id_list_view_width * 1/ 12)
			self.idListView.setColumnWidth(1, id_list_view_width * 1/ 12)
			self.idListView.setColumnWidth(2, id_list_view_width * 9/ 12)
			self.idListView.showGrid()

			self.failedResultModel.setHorizontalHeaderLabels(['ID', 'Status', 'Message'])
			id_list_view_width = self.failedListView.width()
			self.failedListView.setColumnWidth(0, id_list_view_width * 1/ 12)
			self.failedListView.setColumnWidth(1, id_list_view_width * 1/ 12)
			self.failedListView.setColumnWidth(2, id_list_view_width * 9/ 12)
			self.failedListView.showGrid()

			self.threads = []
			for id in self.idList:
				c += 1
				vs = FetchFromAPI(c, id)
				vs.signal.connect(self.show_result)
				self.threads.append(vs)
				vs.start()
	
	def show_result(self, json):
		data = json.get('data')
		if data.get('success'):
			id_model = QStandardItem(data.get('licenseID'))
			success_model = QStandardItem("OK")
			if data.get('is_data_exist'):
				msg_modle = QStandardItem(data.get('message').get('Name'))
			else:
				msg_modle = QStandardItem(data.get('message').get('text'))
			self.resultModel.setItem(self.success_id_count, 0, id_model)
			self.resultModel.setItem(self.success_id_count, 1, success_model)
			self.resultModel.setItem(self.success_id_count, 2, msg_modle)
			self.success_id_count += 1
			self.result_list.append(data)
		else:
			id_model = QStandardItem(data.get('licenseID'))
			success_model = QStandardItem("FAIL")
			msg_modle = QStandardItem(data.get('message').get('errorText'))
			self.failedResultModel.setItem(self.fail_id_count, 0, id_model)
			self.failedResultModel.setItem(self.fail_id_count, 1, success_model)
			self.failedResultModel.setItem(self.fail_id_count, 2, msg_modle)

			self.fail_id_list.append(json.get('data').get('licenseID'))
			self.fail_id_count += 1

		self.passed_id_count += 1
		if self.passed_id_count == self.total_id_count:
			self.startBtn.setEnabled(True)
			self.resumeBtn.setEnabled(True)
			self.exportBtn.setEnabled(True)
			print("Successed : {} in {}".format(self.success_id_count, len(self.idList)))

		progress_value = self.passed_id_count * 100/self.total_id_count
		self.progressBar.setValue(progress_value)

	def exportExcel(self):
		if len(self.result_list):
			exporttime = datetime.now().strftime('%m-%d-%Y-%H-%M-%S')
			self.book = Workbook()
			self.sheet = self.book.active
			self.filename = "License-{}.xlsx".format(exporttime)
			row_count = 1
			self.sheet.cell(row=row_count, column=1).value = "name"
			for data in self.result_list:
				row_count += 1
				self.sheet.cell(row=row_count, column=1).value = data.get('licenseID')
				if data.get('is_data_exist'):
					self.sheet.cell(row=row_count, column=2).value = data.get('message').get('Name')
				else:
					self.sheet.cell(row=row_count, column=2).value = data.get('message').get('text')
					continue
				self.sheet.cell(row=row_count, column=3).value = data.get('message').get('RegStatus')
				self.sheet.cell(row=row_count, column=4).value = data.get('message').get('Date')
				self.sheet.cell(row=row_count, column=5).value = data.get('message').get('Expiration')
				self.sheet.cell(row=row_count, column=6).value = data.get('message').get('City employee')
				self.sheet.cell(row=row_count, column=7).value = data.get('message').get('OfficeAddress')
				self.sheet.cell(row=row_count, column=8).value = data.get('message').get('Status')
				self.sheet.cell(row=row_count, column=9).value = data.get('message').get('Business Phone')
				self.sheet.cell(row=row_count, column=10).value = data.get('message').get('Business 1')
				
				rc = 0
				i = 0
				for ins in data.get('message').get('Insurance'):
					rc = i + row_count
					i += 1
					self.sheet.cell(row=rc, column=11).value = ins.get('type')
					self.sheet.cell(row=rc, column=12).value = ins.get('policy')
					self.sheet.cell(row=rc, column=13).value = ins.get('required')
					self.sheet.cell(row=rc, column=14).value = ins.get('Company')
					self.sheet.cell(row=rc, column=15).value = ins.get('ExpirationDate')
				
				rrc = 0
				i = 0
				for endor in data.get('message').get("Endorsements"):
					rrc = i + row_count
					i += 1
					self.sheet.cell(row=rrc, column=16).value = endor.get('Status')
					self.sheet.cell(row=rrc, column=17).value = endor.get('Type')
				if rc > rrc:
					row_count = rc
				else:
					row_count = rrc


			db_dir = "Output/"
			if not os.path.exists(db_dir):
				os.makedirs(db_dir)
			self.book.save(db_dir + self.filename)
			msg = QMessageBox()
			msg.setIcon(QMessageBox.Information)
			msg.setText("File saved sucessfully")
			msg.setInformativeText("{}".format(self.filename))
			msg.setWindowTitle("Information!")
			msg.exec_()
		else:
			msg = QMessageBox()
			msg.setIcon(QMessageBox.Information)
			msg.setText("No License data!")	
			msg.setWindowTitle("Alert!")
			msg.exec_()		

	def resumeScrape(self):
		if len(self.fail_id_list):

			self.startBtn.setEnabled(False)
			self.resumeBtn.setEnabled(False)
			self.exportBtn.setEnabled(False)
			self.progressBar.setValue(0)

			fail_ids = self.fail_id_list.copy()
			self.fail_id_list.clear()
			
			self.fail_id_count = 0
			self.passed_id_count = 0
			self.total_id_count = len(fail_ids)
			self.failedResultModel.clear()

			self.failedResultModel.setHorizontalHeaderLabels(['ID', 'Status', 'Message'])
			id_list_view_width = self.failedListView.width()
			self.failedListView.setColumnWidth(0, id_list_view_width * 1/ 12)
			self.failedListView.setColumnWidth(1, id_list_view_width * 1/ 12)
			self.failedListView.setColumnWidth(2, id_list_view_width * 9/ 12)
			self.failedListView.showGrid()

			self.threads = []
			c = 0
			for id in fail_ids:
				c += 1
				vs = FetchFromAPI(c, id)
				vs.signal.connect(self.show_result)
				self.threads.append(vs)
				vs.start()


	def get_ids_from_edit(self):
		text = self.idsEdit.toPlainText()
		self.idList.clear()
		self.idsEdit.clear()
		if text:
			list = text.strip().split("\n")
			for id in list:
				for item in id.split():
					try:
						int(item)
						self.idList.append(item)
						self.idsEdit.append(item)
					except:
						continue
			self.total_id_count = len(self.idList)
	
class FetchFromAPI(QThread):
	signal = pyqtSignal("PyQt_PyObject")

	def __init__(self, index, id):
		QThread.__init__(self)

		self.api_url = "http://localhost:5000/v1/json"
		self.data = {
			"id" : id
		}
		self.index = index
		self.id = id

	def run(self):
		try:
			# time.sleep(self.index * 1)
			response = requests.post(self.api_url, json=self.data, timeout=100)
			self.signal.emit(response.json())
		except Exception as e:
			self.signal.emit({
				"data" : {
					'success': False,
					"licenseID" : self.id,
					'message' :{
						"errorText" : "Exceed Max requests {} Retry again".format(repr(e))
					} 
				}
			})
		
	
if __name__ == "__main__":
	app =QApplication(sys.argv)
	window = NYCUI()
	app.exec_()